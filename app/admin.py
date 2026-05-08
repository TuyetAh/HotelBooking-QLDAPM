from flask import session, redirect, request, render_template, url_for
from datetime import datetime
from flask_admin import Admin, AdminIndexView, expose
from flask_admin.contrib.sqla import ModelView
from flask_admin.actions import action
from flask_admin.form import BaseForm
from flask import flash
from sqlalchemy import func
from app import db
from flask import request as req
from app.dao import save_room_images
import os
from flask import current_app
from wtforms.validators import Optional, Length
from flask_admin import BaseView
from wtforms import PasswordField, FileField
from app.models import (
    NguoiDung, ChuKhachSan, KhachSan,
    TienIch, LoaiPhong, DatPhong,
    ChiTietDatPhong, ThanhToan, HoanTien,
    DanhGia, ChuyenTienKhachSan
)


# =========================================================
# BẢO MẬT
# =========================================================
class SecureModelView(ModelView):
    form_base_class = BaseForm

    def is_accessible(self):
        # [SỬA] Dùng session admin riêng thay vì session.get("vai_tro") == 0
        return session.get("admin_logged_in") is True

    def inaccessible_callback(self, name, **kwargs):
        # [SỬA] Redirect về trang login admin (không phải /dang-nhap)
        return redirect("/admin/login")


# =========================================================
# DASHBOARD
# =========================================================
class DashboardView(AdminIndexView):

    # =========================================================
    # [THÊM] _handle_view — bypass is_accessible cho login & logout
    #
    # Vấn đề gốc: Flask-Admin gọi is_accessible() cho MỌI route trong view,
    # kể cả route /login. Khi chưa login → inaccessible_callback redirect về
    # /admin/login → lại gọi is_accessible() → vòng lặp vô tận (ERR_TOO_MANY_REDIRECTS).
    #
    # Giải pháp: override _handle_view, trả về None cho login/logout để
    # Flask-Admin không chặn 2 route đó, các route khác vẫn bị bảo vệ bình thường.
    # =========================================================
    def _handle_view(self, name, **kwargs):
        # Cho phép truy cập tự do vào login và logout (không cần đăng nhập)
        if name in ("login", "logout"):
            return None
        # Các route còn lại: chưa login thì chuyển về trang login
        if not session.get("admin_logged_in"):
            return redirect("/admin/login")
        return None

    @expose("/")
    def index(self):
        stats = {
            "total_users":        NguoiDung.query.count(),
            "total_hotels":       KhachSan.query.count(),
            "total_bookings":     DatPhong.query.count(),
            "pending_bookings":   DatPhong.query.filter_by(TrangThaiDatPhong=0).count(),
            "confirmed_bookings": DatPhong.query.filter_by(TrangThaiDatPhong=1).count(),
            "cancelled_bookings": DatPhong.query.filter_by(TrangThaiDatPhong=2).count(),
            "completed_bookings": DatPhong.query.filter_by(TrangThaiDatPhong=3).count(),
            "total_revenue": db.session.query(
                func.sum(ThanhToan.SoTienThanhToan)
            ).filter_by(TrangThaiThanhToan=1).scalar() or 0,
            "recent_bookings": DatPhong.query.order_by(
                DatPhong.NgayTao.desc()
            ).limit(10).all(),
        }
        return self.render("admin/index.html", stats=stats)

    # =========================================================
    # [THÊM] Route đăng nhập Admin — GET: hiển thị form, POST: xử lý
    # =========================================================
    @expose("/login", methods=["GET", "POST"])
    def login(self):
        # Nếu đã đăng nhập thì vào thẳng dashboard
        if session.get("admin_logged_in"):
            return redirect("/admin/")

        error = None

        if request.method == "POST":
            username = request.form.get("username", "").strip()
            password = request.form.get("password", "").strip()

            from werkzeug.security import check_password_hash
            # Chỉ cho đăng nhập nếu VaiTro == 0 (Admin)
            user = NguoiDung.query.filter_by(
                TenDangNhap=username,
                VaiTro=0
            ).first()

            from werkzeug.security import check_password_hash

            is_valid = False
            try:
                is_valid = check_password_hash(user.MatKhau, password)
            except Exception:
                pass

            # Fallback: plain text (dùng cho data test)
            if not is_valid:
                is_valid = (user.MatKhau == password)

            if user and is_valid:
                session["admin_logged_in"] = True
                session["admin_ho_ten"]    = user.HoTen
                session["admin_username"]  = user.TenDangNhap
                session["vai_tro"]         = 0  # giữ để ThongKeView cũ tương thích
                return redirect("/admin/")
            else:
                error = "Sai tên đăng nhập / mật khẩu, hoặc tài khoản không có quyền admin."

        return self.render("admin/login.html", error=error)

    # =========================================================
    # [THÊM] Route đăng xuất Admin
    # =========================================================
    @expose("/logout")
    def logout(self):
        session.pop("admin_logged_in", None)
        session.pop("admin_ho_ten",    None)
        session.pop("admin_username",  None)
        session.pop("vai_tro",         None)
        return redirect("/admin/login")


# =========================================================
# VIEWS — giữ nguyên logic, chỉ kế thừa SecureModelView đã sửa
# =========================================================
class NguoiDungView(SecureModelView):
    column_list            = ["MaNguoiDung", "TenDangNhap", "HoTen", "Email",
                               "SoDienThoai", "VaiTro", "TrangThaiHoatDong", "NgayTao"]
    column_searchable_list = ["TenDangNhap", "HoTen", "Email"]
    column_filters         = ["VaiTro", "TrangThaiHoatDong"]
    column_labels          = {
        "MaNguoiDung": "Mã", "TenDangNhap": "Tên đăng nhập",
        "HoTen": "Họ tên", "Email": "Email", "SoDienThoai": "SĐT",
        "VaiTro": "Vai trò", "TrangThaiHoatDong": "Trạng thái", "NgayTao": "Ngày tạo"
    }
    form_columns = [
        "TenDangNhap", "mat_khau_nhap", "HoTen", "SoDienThoai",
        "Email", "SoTaiKhoanNganHang", "VaiTro", "TrangThaiHoatDong",
    ]
    form_extra_fields = {
        "mat_khau_nhap": PasswordField("Mật khẩu",
                                       validators=[Optional(), Length(min=6, max=255)])
    }

    def on_model_change(self, form, model, is_created):
        from werkzeug.security import generate_password_hash
        raw_password = form.mat_khau_nhap.data
        if is_created:
            if not raw_password:
                from wtforms.validators import ValidationError
                raise ValidationError("Mật khẩu không được để trống khi tạo mới.")
            model.MatKhau = generate_password_hash(raw_password)
        else:
            if raw_password:
                model.MatKhau = generate_password_hash(raw_password)

    can_export = True
    page_size  = 20


class KhachSanView(SecureModelView):
    def get_query(self):
        return self.session.query(self.model).filter(self.model.TrangThaiDuyet == 1)

    def get_count_query(self):
        return self.session.query(func.count('*')).filter(self.model.TrangThaiDuyet == 1)

    column_list            = ["MaKhachSan", "TenKhachSan", "ThanhPho",
                               "TrangThaiDuyet", "TrangThaiHoatDong",
                               "DiemDanhGiaTrungBinh", "NgayTao"]
    column_searchable_list = ["TenKhachSan", "ThanhPho", "DiaChi"]
    column_filters         = ["TrangThaiHoatDong", "ThanhPho"]
    column_labels          = {
        "MaKhachSan": "Mã", "TenKhachSan": "Tên KS", "ThanhPho": "Thành phố",
        "TrangThaiDuyet": "Duyệt", "TrangThaiHoatDong": "Hoạt động",
        "DiemDanhGiaTrungBinh": "Điểm TB", "NgayTao": "Ngày tạo"
    }
    form_columns = [
        "chu_khach_san", "TenKhachSan", "ThanhPho", "DiaChi",
        "ViTriNoiBat", "SoDienThoaiLienHe", "MoTa", "QuyDinhKhachSan",
        "ChinhSachHuy", "ThuMucAnh", "TrangThaiDuyet",
        "LyDoTuChoi", "TrangThaiHoatDong",
    ]
    can_export = True
    page_size  = 20


class KhachSanChoDuyetView(SecureModelView):
    def get_query(self):
        return self.session.query(self.model).filter(self.model.TrangThaiDuyet == 0)

    def get_count_query(self):
        return self.session.query(func.count('*')).filter(self.model.TrangThaiDuyet == 0)

    column_list   = ["MaKhachSan", "TenKhachSan", "ThanhPho",
                     "DiaChi", "SoDienThoaiLienHe", "xem_anh", "NgayTao"]
    column_labels = {
        "MaKhachSan": "Mã", "TenKhachSan": "Tên KS",
        "ThanhPho": "Thành phố", "DiaChi": "Địa chỉ",
        "SoDienThoaiLienHe": "SĐT", "xem_anh": "Ảnh", "NgayTao": "Ngày gửi"
    }

    def _xem_anh_formatter(view, context, model, name):
        if not model.ThuMucAnh:
            return "Chưa có ảnh"
        import os
        from flask import url_for, current_app
        from markupsafe import Markup
        folder_path = os.path.join(current_app.root_path, "static", "images", model.ThuMucAnh)
        if not os.path.exists(folder_path):
            return "Chưa có ảnh"
        valid_ext = (".jpg", ".jpeg", ".png", ".webp", ".gif")
        files = sorted([f for f in os.listdir(folder_path) if f.lower().endswith(valid_ext)])
        if not files:
            return "Chưa có ảnh"
        imgs = ""
        for f in files[:5]:
            src = url_for("static", filename=f"images/{model.ThuMucAnh}/{f}")
            imgs += f'<img src="{src}" style="width:80px; height:60px; object-fit:cover; border-radius:6px; margin:2px;">'
        return Markup(f'<div style="display:flex; flex-wrap:wrap; gap:4px;">{imgs}</div>')

    column_formatters = {"xem_anh": _xem_anh_formatter}
    can_create = False
    can_delete = False
    can_edit   = False
    can_export = True
    page_size  = 20

    @action("duyet", "✅ Duyệt", "Duyệt các khách sạn đã chọn?")
    def action_duyet(self, ids):
        try:
            count = 0
            for id in ids:
                hotel = KhachSan.query.get(id)
                if hotel:
                    hotel.TrangThaiDuyet = 1
                    hotel.NgayDuyet = datetime.now()
                    hotel.LyDoTuChoi = None
                    count += 1
            db.session.commit()
            flash(f"Đã duyệt {count} khách sạn.", "success")
        except Exception as e:
            db.session.rollback()
            flash(f"Lỗi: {str(e)}", "error")

    @action("tu_choi", "❌ Từ chối", "Từ chối các khách sạn đã chọn?")
    def action_tu_choi(self, ids):
        try:
            count = 0
            for id in ids:
                hotel = KhachSan.query.get(id)
                if hotel:
                    hotel.TrangThaiDuyet = 2
                    hotel.NgayDuyet = None
                    hotel.LyDoTuChoi = "Không đáp ứng yêu cầu của hệ thống"
                    count += 1
            db.session.commit()
            flash(f"Đã từ chối {count} khách sạn.", "success")
        except Exception as e:
            db.session.rollback()
            flash(f"Lỗi: {str(e)}", "error")


class KhachSanTuChoiView(SecureModelView):
    def get_query(self):
        return self.session.query(self.model).filter(self.model.TrangThaiDuyet == 2)

    def get_count_query(self):
        return self.session.query(func.count('*')).filter(self.model.TrangThaiDuyet == 2)

    column_list   = ["MaKhachSan", "TenKhachSan", "ThanhPho", "DiaChi", "LyDoTuChoi", "NgayTao"]
    column_labels = {
        "MaKhachSan": "Mã", "TenKhachSan": "Tên KS",
        "ThanhPho": "Thành phố", "DiaChi": "Địa chỉ",
        "LyDoTuChoi": "Lý do từ chối", "NgayTao": "Ngày gửi"
    }
    can_create = False
    can_delete = False
    can_edit   = True
    can_export = True
    form_columns = ["LyDoTuChoi"]
    page_size  = 20


class LoaiPhongView(SecureModelView):
    column_list            = ["MaLoaiPhong", "khach_san", "TenLoaiPhong",
                               "GiaMoiDem", "SoNguoiToiDa", "SoLuongPhong", "TrangThaiHoatDong"]
    column_searchable_list = ["TenLoaiPhong"]
    column_filters         = ["TrangThaiHoatDong"]
    column_labels          = {
        "khach_san": "Khách sạn", "TenLoaiPhong": "Tên loại phòng",
        "GiaMoiDem": "Giá/đêm", "SoNguoiToiDa": "Số người tối đa",
        "SoLuongPhong": "Số lượng phòng", "TrangThaiHoatDong": "Trạng thái"
    }
    form_columns = [
        "khach_san", "TenLoaiPhong", "MoTa",
        "GiaMoiDem", "SoNguoiToiDa", "SoLuongPhong", "TrangThaiHoatDong",
        "room_images",   # <-- phải có trong form_columns mới hiện ra
    ]
    form_extra_fields = {
        "room_images": FileField("Ảnh loại phòng")
    }
    form_widget_args = {
        "room_images": {"multiple": True}   # cho phép chọn nhiều ảnh
    }
    can_export = True
    page_size  = 20

    def on_form_prefill(self, form, id):
        # Đảm bảo form có enctype multipart khi edit
        pass

    def create_form(self, obj=None):
        form = super().create_form(obj)
        return form

    def after_model_change(self, form, model, is_created):
        # Nếu chưa có ThuMucAnh thì tạo (giống create_room_type trong dao.py)
        if not model.ThuMucAnh:
            model.ThuMucAnh = f"loaiphong/lp_{model.MaLoaiPhong}"
            folder_path = os.path.join(
                current_app.root_path,
                "static", "images",
                model.ThuMucAnh
            )
            os.makedirs(folder_path, exist_ok=True)

            from app import db
            db.session.commit()

        files = req.files.getlist("room_images")
        if files and files[0].filename != "":
            save_room_images(model.ThuMucAnh, files)


class DatPhongView(SecureModelView):
    column_list            = ["MaDatPhong", "MaDatPhongCode", "nguoi_dung",
                               "khach_san", "NgayNhanPhong", "NgayTraPhong",
                               "TongTien", "TrangThaiDatPhong", "NgayTao"]
    column_searchable_list = ["MaDatPhongCode"]
    column_filters         = ["TrangThaiDatPhong"]
    column_labels          = {
        "MaDatPhongCode": "Mã code", "nguoi_dung": "Khách hàng",
        "khach_san": "Khách sạn", "NgayNhanPhong": "Nhận phòng",
        "NgayTraPhong": "Trả phòng", "TongTien": "Tổng tiền",
        "TrangThaiDatPhong": "Trạng thái", "NgayTao": "Ngày đặt"
    }
    form_columns = [
        "MaDatPhongCode", "nguoi_dung", "khach_san",
        "NgayNhanPhong", "NgayTraPhong",
        "SoNguoiLuuTru", "TongTien", "TrangThaiDatPhong",
    ]
    can_delete = False
    can_export = True
    page_size  = 20


class ThanhToanView(SecureModelView):
    column_list    = ["MaThanhToan", "dat_phong", "PhuongThucThanhToan",
                      "TrangThaiThanhToan", "SoTienThanhToan", "ThoiGianThanhToan"]
    column_filters = ["TrangThaiThanhToan", "PhuongThucThanhToan"]
    column_labels  = {
        "dat_phong": "Đơn đặt", "PhuongThucThanhToan": "Phương thức",
        "TrangThaiThanhToan": "Trạng thái", "SoTienThanhToan": "Số tiền",
        "ThoiGianThanhToan": "Thời gian"
    }
    form_columns = [
        "dat_phong", "PhuongThucThanhToan", "TrangThaiThanhToan",
        "SoTienThanhToan", "MaGiaoDich", "ThoiGianThanhToan",
    ]
    can_delete = False
    can_export = True
    page_size  = 20


class DanhGiaView(SecureModelView):
    column_list            = ["MaDanhGia", "nguoi_dung", "khach_san",
                               "SoSao", "BinhLuan", "NgayDanhGia"]
    column_searchable_list = ["BinhLuan"]
    column_filters         = ["SoSao"]
    column_labels          = {
        "nguoi_dung": "Khách hàng", "khach_san": "Khách sạn",
        "SoSao": "Số sao", "BinhLuan": "Bình luận", "NgayDanhGia": "Ngày đánh giá"
    }
    form_columns = ["dat_phong", "nguoi_dung", "khach_san", "SoSao", "BinhLuan"]
    can_export = True
    page_size  = 20


class HoanTienView(SecureModelView):
    column_list   = ["MaHoanTien", "dat_phong", "SoTienHoan",
                     "LyDoHoanTien", "TrangThaiHoanTien", "ThoiGianHoanTien"]
    column_labels = {
        "dat_phong": "Đơn đặt", "SoTienHoan": "Số tiền hoàn",
        "LyDoHoanTien": "Lý do", "TrangThaiHoanTien": "Trạng thái",
        "ThoiGianHoanTien": "Thời gian"
    }
    form_columns = [
        "dat_phong", "SoTienHoan", "LyDoHoanTien", "TrangThaiHoanTien", "ThoiGianHoanTien",
    ]
    can_export = True
    page_size  = 20


class ChuyenTienView(SecureModelView):
    column_list   = ["MaChuyenTien", "dat_phong", "khach_san",
                     "TongTienDonHang", "PhiHeThong",
                     "SoTienChuyenChoKhachSan", "TrangThaiChuyenTien"]
    column_labels = {
        "dat_phong": "Đơn đặt", "khach_san": "Khách sạn",
        "TongTienDonHang": "Tổng đơn", "PhiHeThong": "Phí hệ thống",
        "SoTienChuyenChoKhachSan": "Tiền chuyển KS", "TrangThaiChuyenTien": "Trạng thái"
    }
    form_columns = [
        "dat_phong", "khach_san", "TongTienDonHang", "PhiHeThong",
        "SoTienChuyenChoKhachSan", "TrangThaiChuyenTien", "ThoiGianChuyenTien",
    ]
    can_export = True
    page_size  = 20


class ThongKeView(BaseView):

    # [THÊM] Bảo vệ ThongKeView bằng session admin
    def is_accessible(self):
        return session.get("admin_logged_in") is True

    def inaccessible_callback(self, name, **kwargs):
        return redirect("/admin/login")

    @expose("/")
    def index(self):
        if not session.get("admin_logged_in"):
            return redirect("/admin/login")

        from app.models import KhachSan, ThanhToan, HoanTien, ChuyenTienKhachSan, DatPhong
        from sqlalchemy import func, extract
        import json

        khach_sans = KhachSan.query.filter_by(TrangThaiDuyet=1).all()
        thong_ke_ks = []
        for ks in khach_sans:
            tong_doanh_thu = db.session.query(func.sum(ThanhToan.SoTienThanhToan)).join(DatPhong).filter(
                DatPhong.MaKhachSan == ks.MaKhachSan, ThanhToan.TrangThaiThanhToan == 1).scalar() or 0
            can_chuyen = db.session.query(func.sum(ChuyenTienKhachSan.SoTienChuyenChoKhachSan)).filter(
                ChuyenTienKhachSan.MaKhachSan == ks.MaKhachSan, ChuyenTienKhachSan.TrangThaiChuyenTien == 0).scalar() or 0
            da_chuyen = db.session.query(func.sum(ChuyenTienKhachSan.SoTienChuyenChoKhachSan)).filter(
                ChuyenTienKhachSan.MaKhachSan == ks.MaKhachSan, ChuyenTienKhachSan.TrangThaiChuyenTien == 1).scalar() or 0
            phi_he_thong = db.session.query(func.sum(ChuyenTienKhachSan.PhiHeThong)).filter(
                ChuyenTienKhachSan.MaKhachSan == ks.MaKhachSan, ChuyenTienKhachSan.TrangThaiChuyenTien == 1).scalar() or 0
            can_hoan = db.session.query(func.sum(HoanTien.SoTienHoan)).join(DatPhong).filter(
                DatPhong.MaKhachSan == ks.MaKhachSan, HoanTien.TrangThaiHoanTien == 0).scalar() or 0
            da_hoan = db.session.query(func.sum(HoanTien.SoTienHoan)).join(DatPhong).filter(
                DatPhong.MaKhachSan == ks.MaKhachSan, HoanTien.TrangThaiHoanTien == 1).scalar() or 0
            thong_ke_ks.append({
                "ks": ks, "tong_doanh_thu": tong_doanh_thu,
                "can_chuyen": can_chuyen, "da_chuyen": da_chuyen,
                "phi_he_thong": phi_he_thong, "can_hoan": can_hoan, "da_hoan": da_hoan,
            })

        from datetime import datetime, date
        nam_hien_tai = datetime.now().year
        doanh_thu_thang = []
        for thang in range(1, 13):
            dt = db.session.query(func.sum(ThanhToan.SoTienThanhToan)).filter(
                ThanhToan.TrangThaiThanhToan == 1,
                extract('year', ThanhToan.ThoiGianThanhToan) == nam_hien_tai,
                extract('month', ThanhToan.ThoiGianThanhToan) == thang
            ).scalar() or 0
            doanh_thu_thang.append(float(dt))

        doanh_thu_quy = []
        for quy in range(1, 5):
            thang_dau = (quy - 1) * 3 + 1
            thang_cuoi = quy * 3
            dt = db.session.query(func.sum(ThanhToan.SoTienThanhToan)).filter(
                ThanhToan.TrangThaiThanhToan == 1,
                extract('year', ThanhToan.ThoiGianThanhToan) == nam_hien_tai,
                extract('month', ThanhToan.ThoiGianThanhToan) >= thang_dau,
                extract('month', ThanhToan.ThoiGianThanhToan) <= thang_cuoi
            ).scalar() or 0
            doanh_thu_quy.append(float(dt))

        doanh_thu_nam = []
        labels_nam = []
        for nam in range(nam_hien_tai - 4, nam_hien_tai + 1):
            dt = db.session.query(func.sum(ThanhToan.SoTienThanhToan)).filter(
                ThanhToan.TrangThaiThanhToan == 1,
                extract('year', ThanhToan.ThoiGianThanhToan) == nam
            ).scalar() or 0
            doanh_thu_nam.append(float(dt))
            labels_nam.append(str(nam))

        tong_phi_he_thong = db.session.query(func.sum(ChuyenTienKhachSan.PhiHeThong)).filter(
            ChuyenTienKhachSan.TrangThaiChuyenTien == 1).scalar() or 0

        return self.render(
            "admin/thong_ke.html",
            thong_ke_ks=thong_ke_ks,
            doanh_thu_thang=json.dumps(doanh_thu_thang),
            doanh_thu_quy=json.dumps(doanh_thu_quy),
            doanh_thu_nam=json.dumps(doanh_thu_nam),
            labels_nam=json.dumps(labels_nam),
            nam_hien_tai=nam_hien_tai,
            tong_phi_he_thong=tong_phi_he_thong
        )

    @expose("/export")
    def export(self):
        if not session.get("admin_logged_in"):
            return redirect("/admin/login")

        from app.models import KhachSan, ThanhToan, HoanTien, ChuyenTienKhachSan, DatPhong
        from sqlalchemy import func
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from flask import request, send_file
        from datetime import datetime

        ma_ks_list = request.args.getlist("ma_ks", type=int)
        if ma_ks_list:
            khach_sans = KhachSan.query.filter(
                KhachSan.MaKhachSan.in_(ma_ks_list), KhachSan.TrangThaiDuyet == 1).all()
        else:
            khach_sans = KhachSan.query.filter_by(TrangThaiDuyet=1).all()

        rows = []
        for ks in khach_sans:
            tong_doanh_thu = db.session.query(func.sum(ThanhToan.SoTienThanhToan)).join(DatPhong).filter(
                DatPhong.MaKhachSan == ks.MaKhachSan, ThanhToan.TrangThaiThanhToan == 1).scalar() or 0
            can_chuyen = db.session.query(func.sum(ChuyenTienKhachSan.SoTienChuyenChoKhachSan)).filter(
                ChuyenTienKhachSan.MaKhachSan == ks.MaKhachSan, ChuyenTienKhachSan.TrangThaiChuyenTien == 0).scalar() or 0
            da_chuyen = db.session.query(func.sum(ChuyenTienKhachSan.SoTienChuyenChoKhachSan)).filter(
                ChuyenTienKhachSan.MaKhachSan == ks.MaKhachSan, ChuyenTienKhachSan.TrangThaiChuyenTien == 1).scalar() or 0
            phi_he_thong = db.session.query(func.sum(ChuyenTienKhachSan.PhiHeThong)).filter(
                ChuyenTienKhachSan.MaKhachSan == ks.MaKhachSan, ChuyenTienKhachSan.TrangThaiChuyenTien == 1).scalar() or 0
            can_hoan = db.session.query(func.sum(HoanTien.SoTienHoan)).join(DatPhong).filter(
                DatPhong.MaKhachSan == ks.MaKhachSan, HoanTien.TrangThaiHoanTien == 0).scalar() or 0
            da_hoan = db.session.query(func.sum(HoanTien.SoTienHoan)).join(DatPhong).filter(
                DatPhong.MaKhachSan == ks.MaKhachSan, HoanTien.TrangThaiHoanTien == 1).scalar() or 0
            rows.append([ks.TenKhachSan, ks.ThanhPho, tong_doanh_thu,
                         can_chuyen, da_chuyen, phi_he_thong, can_hoan, da_hoan])

        wb = Workbook()
        ws = wb.active
        ws.title = "Thống kê tài chính"
        header_fill = PatternFill("solid", start_color="0D6EFD", end_color="0D6EFD")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        total_fill  = PatternFill("solid", start_color="F0F4FF", end_color="F0F4FF")
        total_font  = Font(bold=True, size=11)
        thin   = Side(style="thin", color="D1D5DB")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        center = Alignment(horizontal="center", vertical="center")
        vnd_fmt = '#,##0 "đ"'

        ws.merge_cells("A1:H1")
        ws["A1"] = f"THỐNG KÊ TÀI CHÍNH KHÁCH SẠN — Xuất ngày {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws["A1"].font = Font(bold=True, size=14, color="111827")
        ws["A1"].alignment = center
        ws.row_dimensions[1].height = 36

        headers = ["Khách sạn", "Thành phố", "Tổng doanh thu", "Cần chuyển owner",
                   "Đã chuyển owner", "Phí hệ thống", "Cần hoàn tiền", "Đã hoàn tiền"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=h)
            cell.font = header_font; cell.fill = header_fill
            cell.alignment = center; cell.border = border
        ws.row_dimensions[2].height = 28

        for r_idx, row in enumerate(rows, 3):
            for c_idx, val in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.border = border
                cell.alignment = Alignment(vertical="center",
                                           horizontal="right" if c_idx > 2 else "left")
                if c_idx > 2:
                    cell.number_format = vnd_fmt
            ws.row_dimensions[r_idx].height = 22

        last_data_row = 2 + len(rows)
        total_row = last_data_row + 1
        ws.cell(row=total_row, column=1, value="TỔNG CỘNG").font = total_font
        ws.cell(row=total_row, column=1).fill = total_fill
        ws.cell(row=total_row, column=1).border = border
        for col in range(2, 9):
            cell = ws.cell(row=total_row, column=col)
            cell.value = "" if col == 2 else f"=SUM({get_column_letter(col)}3:{get_column_letter(col)}{last_data_row})"
            if col != 2:
                cell.number_format = vnd_fmt
            cell.font = total_font; cell.fill = total_fill; cell.border = border
            cell.alignment = Alignment(horizontal="right", vertical="center")
        ws.row_dimensions[total_row].height = 26

        col_widths = [32, 18, 20, 22, 20, 18, 18, 18]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"thong_ke_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return send_file(output,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name=filename)


# =========================================================
# KHỞI TẠO ADMIN — giữ nguyên hoàn toàn
# =========================================================
def init_admin(app):
    admin = Admin(
        app,
        name="Hotel Booking Admin",
        index_view=DashboardView(name="Dashboard", url="/admin")
    )
    admin.add_view(NguoiDungView(NguoiDung, db.session,
                                 name="Người dùng", category="Quản lý"))
    admin.add_view(KhachSanChoDuyetView(KhachSan, db.session,
                                        name="Chờ duyệt",
                                        endpoint="khachsan_cho_duyet",
                                        category="Duyệt KS"))
    admin.add_view(KhachSanTuChoiView(KhachSan, db.session,
                                      name="Đã từ chối",
                                      endpoint="khachsan_tu_choi",
                                      category="Duyệt KS"))
    admin.add_view(KhachSanView(KhachSan, db.session,
                                name="Tất cả KS",
                                endpoint="khachsan_tat_ca",
                                category="Duyệt KS"))
    admin.add_view(LoaiPhongView(LoaiPhong, db.session,
                                 name="Loại phòng", category="Quản lý"))
    admin.add_view(DatPhongView(DatPhong, db.session,
                                name="Đặt phòng", category="Quản lý"))
    admin.add_view(ThanhToanView(ThanhToan, db.session,
                                 name="Thanh toán", category="Quản lý"))
    admin.add_view(HoanTienView(HoanTien, db.session,
                                name="Hoàn tiền", category="Quản lý"))
    admin.add_view(DanhGiaView(DanhGia, db.session,
                               name="Đánh giá", category="Quản lý"))
    admin.add_view(ChuyenTienView(ChuyenTienKhachSan, db.session,
                                  name="Chuyển tiền", category="Quản lý"))
    admin.add_view(SecureModelView(TienIch, db.session,
                                   name="Tiện ích", category="Cấu hình"))
    admin.add_view(ThongKeView(name="📊 Thống kê", url="thong-ke", endpoint="thong_ke"))
    return admin